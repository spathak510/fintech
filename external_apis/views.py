from rest_framework.response import Response
from rest_framework.views import APIView
import requests, json, re, os, logging,inspect, traceback
from external_apis.models import GstVerification, TempFiles, DocumentTampering
from external_apis.serializers import GstVerificationSerializer, McaVerificationSerializer
from datetime import date, datetime
from hv_whatsapp import settings as app_settings
from rest_framework import generics, status, permissions
from external_apis.utils import KarzaFunction, HitSaver, UanHandler, detect_text, margedict, encode_credentials
from external_apis.passport_regex import get_indian_passport_front_result, get_indian_passport_back_result
from external_apis.pancard_regex import get_pan_card_ocr_json_result
from external_apis.voter_id_regex import get_voter_id_front_json, get_voter_id_back_json
from external_apis.dl_regex import get_dl_front_json, get_dl_back_json
from external_apis.aadhaar_regex import get_aadhaar_front_json, get_aadhaar_back_json
from django.http import HttpResponse

# Configure logging
logging.basicConfig(filename='external_api.log')




class VerifyGst(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        gst_number = request.GET.get("gst_number")
        if gst_number and len(gst_number) == 15:
            response1 = requests.request("GET", app_settings.EXTERNAL_API['GST_URL'].format(gst_number=gst_number))
            if int(date.today().month)<=3:
                financial_year = str(int(date.today().year)-1)+"-"+str(int(date.today().year))[2:]
            else:
                financial_year = str(int(date.today().year))+"-"+str(int(date.today().year)+1)[2:]
                        
            response = requests.request("GET", app_settings.EXTERNAL_API['GST_DETAIL_URL'].format(gst_number=gst_number,financial_year=financial_year))
            raw_resp = {}
            raw_resp["details"] = json.loads(response1.text)
            raw_resp["filing_history"] = json.loads(response.text)
            
            try:
                gst_status = eval((response1.text).replace('true',"True"))["data"]["sts"]
                last_tax_paid = json.loads(response.text).get("data").get("EFiledlist")[0].get("dof")
            except Exception:
                return Response(data={"status": False,"message":"We are not able to find details of this gst number!"},status=400)
            final_gst = {"gst_number":gst_number,"gst_status":gst_status,"last_tax_paid_date":last_tax_paid,"raw_response":str(raw_resp)}
            gst_serialized = GstVerificationSerializer(data=final_gst)
            if gst_serialized.is_valid():
                gst_serialized.save()
            else:
                return Response(data={"data":gst_serialized.errors,"status": False,"message":"Their is some error from our side!"},status=400)        
            return Response(data={"data":gst_serialized.data,"status": True,"message":"Gst Number Check successfully!"},status=200)
        return Response(data={"status": False,"message":"Invalid Gst Number!"},status=400)


class VerifyMca(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        company_name = request.GET.get("company_name")
        if company_name:
            response1 = requests.request("GET", app_settings.EXTERNAL_API['MCA_URL'].format(company_name=company_name))
            company_basic_detail = json.loads(response1.text)
            raw_resp = {}
            try:
                raw_resp = company_basic_detail[0]
                raw_resp["address"] = raw_resp.get("state")
                cin_number = raw_resp.get("ref")
                company_name = raw_resp.get("name")
                status = raw_resp.get("status")
                address = raw_resp.get("address")
                industry = raw_resp.get("industry")
                incorporation_date = raw_resp.get("incorporation_date")[0:10]
                directors = raw_resp.get("directors")
                for director in directors:
                    director["din"] = director.get("slug")[10:18]
                    del director["slug"]
            except Exception:
                return Response(data={"status": False,"message":"We are not able to find details of this mca!"},status=400)
            final_mca = {"cin_number":cin_number,"company_name":company_name,"status":status,"address":address,"industry":industry,"incorporation_date":incorporation_date,"directors":str(directors),"raw_response":str(raw_resp)}
            mca_serialized = McaVerificationSerializer(data=final_mca)
            if mca_serialized.is_valid():
                mca_serialized.save()
            else:
                return Response(data={"data":mca_serialized.errors,"status": False,"message":"Their is some error from our side!"},status=400)        
            return Response(data={"data":mca_serialized.data,"status": True,"message":"mca Check successfully!"},status=200)
        return Response(data={"status": False,"message":"Company name required!"},status=400)

class EmailDomainVerification(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self,request):
        try:
            pattern = "^\S+@\S+\.\S+$"
            objs = re.search(pattern, request.data.get("email"))
            try:
                if objs.string == request.data.get("email"):
                    pass
            except Exception:
                return Response(data={"status": False,"message":"Invalid email provided!"},status=400)
            payload = json.dumps({
            "email": request.data.get("email")
            })

            url=app_settings.EXTERNAL_API['EMAIL_URL']
            headers = {
            app_settings.EXTERNAL_API['KARZA_API_KEY_ID']: app_settings.EXTERNAL_API['KARZAAPI_KEY'],
            'Content-Type': 'application/json'
            }

            response = requests.request("POST", url, headers=headers, data=payload)
            return Response(data={"status": True,"message":json.loads(response.text)},status=200)
        except Exception:
            return Response(data={"status": False,"message":"API down!"},status=400)

class UanVerification(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self,request):
        try:
            url=app_settings.EXTERNAL_API['UAN_URL']

            payload = json.dumps({
            "uans": [
                request.data.get("uan")
            ],
            "entityId": request.data.get("entityId"),
            "employerName": request.data.get("employerName"),
            "employeeName": request.data.get("employeeName"),
            "mobile": request.data.get("mobile"),
            "emailId": request.data.get("emailId"),
            "pdf": True
            })
            # return Response(data={"status": False,"message":json.loads(payload)},status=400)
            headers = {
            app_settings.EXTERNAL_API['KARZA_API_KEY_ID']: app_settings.EXTERNAL_API['KARZAAPI_KEY_ADVANCED'],
            'Content-Type': 'application/json'
            }
            
            # return Response(data={"status": False,"message":json.loads(headers)},status=200)
            
            try:
                response = requests.request("POST", url, headers=headers, data=payload)
                
                return Response(data={"status": True,"message":json.loads(response.text)},status=200)
            except Exception:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
        except Exception:
            return Response(data={"status": False,"message":"API down!"},status=400)

class LexisNexis(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            lexis_token_url = app_settings.EXTERNAL_API['LEXUS_TOKEN_URL']
            
            user=app_settings.EXTERNAL_API['LEXUS_TOKEN_USER']
            password=app_settings.EXTERNAL_API['LEXUS_TOKEN_PASSWORD']
            resp = requests.post(lexis_token_url, data={}, auth=(user, password))
            
            try:
                resp= ((resp).json())['access_token']
            except Exception:
                return Response(data={"status": False,"message":"Authentication Error"},status=401)
            
            url = app_settings.EXTERNAL_API['LEXUS_MAIN_URL']
            data=[
            request.data.get("client_reference"),request.data.get("entity_type"),
            request.data.get("gender"),request.data.get("first_name"),
            request.data.get("middle_name"),request.data.get("last_name"),
            request.data.get("city"),request.data.get("postal_code"),
            request.data.get("country"),request.data.get("DOB"),
            request.data.get("doc_no"),request.data.get("doc_type")
            ]
            conv = lambda i : i or None
            data = [conv(i) for i in data]
            
            # return Response(data={"status": False,"message":json.loads(data)},status=400)
            try:
                from datetime import datetime
                now = datetime.now()
                payload=json.dumps({
                "ClientContext": {
                    "ClientReference": data[0]
                },
                "SearchConfiguration": {
                    "AssignResultTo": {
                    "Division": "Default Division",
                    "EmailNotification": True,
                    "RolesOrUsers": [
                        "Administrator"
                    ],
                    "Type": "Role"
                    },
                    "PredefinedSearchName": "List Screening",
                    "WriteResultsToDatabase": False
                },
                "SearchInput": {
                    "Records": [
                    {
                        "Entity": {
                        "EntityType": data[1],
                        "Gender": data[2],
                        "Name": {
                            "First": data[3],
                            "Middle": data[4],
                            "Last": data[5]
                        },
                        "Addresses": [
                            {
                            "City": data[6],
                            "PostalCode": data[7],
                            "Country": data[8],
                            "Type": "Current"
                            }
                        ],
                        "AdditionalInfo": [
                            {
                            "Date": {
                                "Day": int(now.strftime("%d")),
                                "Month": int(now.strftime("%m")),
                                "Year": int(now.strftime("%Y"))
                            },
                            "Label": "DateOfBirth",
                            "Type": "DOB",
                            "Value": data[9]
                            }
                        ]
                        },
                        "ID": {
                        "Number": data[10],
                        "Type": data[11]
                        }
                    }
                    ]
                }
                })
                
            except Exception:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
            # payload = json.dumps(request.data.get("provided_json"))
            headers = {
            app_settings.EXTERNAL_API['X_API_KEY']:app_settings.EXTERNAL_API['X_API_KEY_VALUE'],
            'Authorization': 'Bearer ' + resp,
            'Content-Type': 'application/json'
            }
            
            response = requests.request("POST", url, headers=headers, data=payload)
            return Response(data={"status": True,"message":json.loads(response.text)},status=200)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
class PanCustom(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self,request):
        url = app_settings.EXTERNAL_API['KARZAAPI_PAN']
        headers = {
            app_settings.EXTERNAL_API['KARZA_API_KEY_ID']:app_settings.EXTERNAL_API['KARZAAPI_KEY'],
            'Content-Type': 'application/json'
            
        }
        payload = json.dumps({
        "pan": request.data.get('pan_no'),
        "consent": "Y"
        })
        response = requests.request("POST", url, headers=headers, data=payload)
        
        if response.status_code == 200:
            try:
                return Response(data={"status": True,"message":"Name: "+(json.loads(response.text)).get("result").get("name")},status=200)
            except Exception:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
        
class MobileVerificationWithoutOTP(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self,request):
        try:
            url = app_settings.EXTERNAL_API['MOBILE_VERIFICATION']
            mobile_no=request.data.get("mobile_no")
            
            if len(mobile_no) !=10:
                return Response(data={"status": False,"message":"Mobile no. should be exactly 10 digits long."},status=400)
            payload = json.dumps({
            "countryCode": "91",
            "mobile": mobile_no,
            "consent": "Y"
            })
            headers = {
            app_settings.EXTERNAL_API['KARZA_API_KEY_ID']:app_settings.EXTERNAL_API['KARZAAPI_KEY_ADVANCED'],
            'Content-Type': 'application/json'
            }
            
            response = requests.request("POST", url, headers=headers, data=payload)
            try:
                return Response(data={"status": True,"message":json.loads(response.text)},status=200)
            except Exception:
                return Response(data={"status": True,"message":"Data not provided properly!"},status=200)
            
        except Exception:
            return Response(data={"status": True,"message":"API down!"},status=200)
        
class LocationToLatLong(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self,request):
        try:
            map_url = 'https://maps.googleapis.com/maps/api/geocode/json?address='
            map_key = 'AIzaSyDl-WGMy2OUgLDCsxZRf5-m28GVJrr2RSw'
            
            req_url = map_url + request.data.get('claimed_address') + '&key=' + map_key
            res = requests.get(req_url)

            location_json = res.json()
            claimed_lat_long = []
            for i in range(len(location_json['results'])):
                lat = location_json['results'][i]['geometry']['location']['lat']
                lng = location_json['results'][i]['geometry']['location']['lng']
                claimed_lat_long.append(str(lat) + ',' + str(lng))
                try:
                    lat = location_json['results'][i]['geometry']['bounds']['northeast']['lat']
                    lng = location_json['results'][i]['geometry']['bounds']['northeast']['lng']
                    claimed_lat_long.append(str(lat) + ',' + str(lng))
                    lat = location_json['results'][i]['geometry']['bounds']['southwest']['lat']
                    lng = location_json['results'][i]['geometry']['bounds']['southwest']['lng']
                    claimed_lat_long.append(str(lat) + ',' + str(lng))
                except Exception:
                    pass
                try:
                    lat = location_json['results'][i]['geometry']["viewport"]['northeast']['lat']
                    lng = location_json['results'][i]['geometry']["viewport"]['northeast']['lng']
                    claimed_lat_long.append(str(lat) + ',' + str(lng))
                    lat = location_json['results'][i]['geometry']["viewport"]['southwest']['lat']
                    lng = location_json['results'][i]['geometry']["viewport"]['southwest']['lng']
                    claimed_lat_long.append(str(lat) + ',' + str(lng))
                except Exception:
                    pass
            lat_long={"lat":lat, "lng":lng}
            return Response(data={"status": True,"message":lat_long},status=200) 
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)

class CompareLatLong(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self,request):
        try:
            import geopy.distance

            coords_1 = (request.data.get('claimed_lat'),request.data.get('claimed_long'))
            coords_2 = (request.data.get('actual_lat'),request.data.get('actual_long'))

            if (geopy.distance.geodesic(coords_1, coords_2).meters) <=300:
                return Response(data={"status": True,"message":{"distance":str(round((geopy.distance.geodesic(coords_1, coords_2).meters),2))+" Meters", "matched":"Yes"}},status=200)
            else:
                return Response(data={"status": True,"message":{"distance":str(round((geopy.distance.geodesic(coords_1, coords_2).meters),2))+" Meters", "matched":"No"}},status=200)
            # return final_claimed_lat_long, request.data.get('actual_lat_long'), min_distance, location_match
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
# score me api starts here:

class CreditRateController(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self,request):
        try:

            payload = json.dumps({
            "entityName": request.data.get('enity_name')
            })
            
            if request.data.get('enity_name')=='' or request.data.get('enity_name')==None:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
            
            if request.data.get("api")=="prod":
                url = app_settings.EXTERNAL_API['SCORE_ME_CREDIT_RATE_CONTROLLER_INPUT_PROD']

                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD'],
                'Content-Type': 'application/json'
                }
            else:
                url = app_settings.EXTERNAL_API['SCORE_ME_CREDIT_RATE_CONTROLLER_INPUT']

                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET'],
                'Content-Type': 'application/json'
                }

            response = requests.request("POST", url, headers=headers, data=payload)
            response = response.json()
            if response.get("responseCode")=="SRS016":
                response= {'referenceId':(response.get('data')).get('referenceId'), 'responseMessage':response.get('responseMessage'),'responseCode':response.get('responseCode')}
                return Response(data={"status": True,"message":response},status=200)
            else:
                return Response(data={"status": False,"message":response},status=400)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
    
    def get(self,request):
        try:
            ref_id=request.GET.get("reference_id")
            
            if ref_id=='' or ref_id==None:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
            payload = ""
            if request.GET.get("api")=="prod":
                url=app_settings.EXTERNAL_API['SCORE_ME_CREDIT_RATE_CONTROLLER_RESULT_PROD'] + ref_id
                
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                }
            else:
                url=app_settings.EXTERNAL_API['SCORE_ME_CREDIT_RATE_CONTROLLER_RESULT'] + ref_id
            
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET']
                }

            response = requests.request("GET", url, headers=headers, data=payload)
            if response.get("responseCode")=="SRC001":
                return Response(data={"status": True,"message":json.loads(response.text)},status=200)
            else:
                return Response(data={"status": False,"message":json.loads(response.text)},status=400)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class CompanyInfo(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self,request):
        try:
            payload = json.dumps({
            "cin_llpin": request.GET.get('cin')
            })

            if request.GET.get('cin')=='' or request.GET.get('cin')==None:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
            
            if request.GET.get("api")=="prod":
                url = app_settings.EXTERNAL_API['SCORE_ME_COMPANY_INFO_PROD']
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD'],
                    'Content-Type': 'application/json'
                    }
            else:
                url = app_settings.EXTERNAL_API['SCORE_ME_COMPANY_INFO']
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET'],
                    'Content-Type': 'application/json'
                    }

            response = requests.request("POST", url, headers=headers, data=payload)
            response = response.json()
            if response.get("responseCode")=="SRC001":
            
                if request.data.get('data')=="all":
                    return Response(data={"status": True,"message":response},status=200)
                else:
                    return Response(data={"status": True,"message":(response['data']['mangementDetails'])},status=200)
            else:
                return Response(data={"status": False,"message":response},status=400)
            
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
    
    # surepass aadhaar card api
class AadhaarCard(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            url = app_settings.EXTERNAL_API['SUREPASS_GENERATE_OTP']

            payload = json.dumps({
            "id_number": request.data.get('aadhaar_num'),
            })
            headers = {
            'Authorization': app_settings.EXTERNAL_API['SUREPASS_AUTH_TOKEN'],
            'Content-Type': 'application/json'
            }

            response = requests.request("POST", url, headers=headers, data=payload)
            # response = ((json.loads(response.text))['data'])
            return Response(data={"status": True,"message":((json.loads(response.text)))},status=200)
                
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
    def get(self, request):
        try:
            url = app_settings.EXTERNAL_API['SUREPASS_SUBMIT_OTP']

            payload = json.dumps({
            "client_id": request.data.get('client_id'),
            "otp": request.data.get('otp')
            })
            headers = {
            'Authorization': app_settings.EXTERNAL_API['SUREPASS_AUTH_TOKEN'],
            'Content-Type': 'application/json'
            }
            response = requests.request("POST", url, headers=headers, data=payload)
            # response = ((json.loads(response.text))['data'])
            return Response(data={"status": True,"message":((json.loads(response.text))['data'])},status=200)
    
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
# score me bank statement
class BankStatementVerification(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            payload={'data': request.data.get('data')}
            if payload['data']==None or payload['data']=='':
                return Response(data={"status": False,"message":"Parameter 'data' cannot be empty."},status=400)
            
            if request.data.get("api")=="prod":
                url = app_settings.EXTERNAL_API['BANK_STATEMENT_UPLOAD_PROD']
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                }
            else:
                url = app_settings.EXTERNAL_API['BANK_STATEMENT_UPLOAD']
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET']
                }

            files = request.FILES.getlist('file')
            if len(files)==1:
                files = request.FILES['file']
                files = {'file': files}
                response = requests.request("POST", url, headers=headers, data=payload, files=files)
                response= response.json()
            elif len(files)>1:
                data = {}
                for i, file in enumerate(files):
                    if i ==0:
                        data[f'file'] = file
                    else:
                        data[f'file{i}'] = file
                response = requests.request("POST", url, headers=headers, data=payload, files=data)
                response = response.json()
            elif len(files)==0:
                return Response(data={"status": False,"message":"File field cannot be empty."},status=400)
            
            try:
                if response.get("responseCode")=="SRS016":
                    return Response(data={"status": True,"message":response},status=200)
                else:
                    return Response(data={"status": False,"message":response},status=400)
            except Exception:
                return Response(data={"status": False,"message":response},status=400)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
    def get(self, request):
        try:
            ref_id = request.GET.get("referenceId")
            user = request.GET.get("user")
            if user not in app_settings.api_users:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
            payload={}

            if request.GET.get("api")=="prod":
                url = app_settings.EXTERNAL_API['BANK_STATEMENT_REPORT_PROD'] + request.GET.get("referenceId")
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                }
            else:
                url = app_settings.EXTERNAL_API['BANK_STATEMENT_REPORT'] + request.GET.get("referenceId")
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET']
                }

            response = requests.request("GET", url, headers=headers, data=payload)
            response = json.loads(response.text)
            data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['BANK_STATEMENT_REPORT'],"payload":ref_id,"response":response,"api_name":"Bank Statement Report - GET"}
            HitSaver(data)
            if response.get("responseCode")=="SRC001":
                if request.GET.get("data_type")=="excel":
                    url =response['data']['excelUrl'] 
                    response = requests.request("GET", url, headers=headers, data=payload)
                    filename = ref_id + ".xlsx"

                    with open(filename, "wb") as file:
                        file.write(response.content)
                    from openpyxl import load_workbook

                    excel_file_path = filename

                    workbook = load_workbook(excel_file_path)

                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        images_to_remove = []
                        for image in sheet._images:
                            images_to_remove.append(image)

                        for image in images_to_remove:
                            sheet._images.remove(image)
                    workbook.save(ref_id + ".xlsx")
                    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    response["Content-Disposition"] = "attachment; filename={}".format(ref_id)
                    with open(ref_id + ".xlsx", "rb") as file:
                        response.write(file.read())
                    os.remove(ref_id + ".xlsx")
                    return response

                else:
                    report_url=response.get('data').get('jsonUrl')
                    response = requests.request("GET",report_url, headers=headers)
                    return Response(data={"status": True,"message":(json.loads(response.text))},status=200)
            # elif response.get("responseCode")=="EIF009":
            #     return Response(data={"status": False,"message":{"message":"Incorrect File","response_code":"EIF009"}},status=400)
            # elif response.get("responseCode")=="RNP020":
            #     return Response(data={"status": False,"message":{"message":"Request under process","response_code":"RNP020"}},status=400)
            # elif response.get("message")=="Credits not available.":
            #     return Response(data={"status": False,"message":{"message":"Unable to process due to technical issues.","response_code":"400"}},status=400)
            else:
                return Response(data={"status":False,"message":response.json()}, status=400)

        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class CreditBalanceSheet(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            payload = {
                'payload': request.data.get('payload')
            }
            if payload['payload']=='':
                return Response(data={"status": False,"message":"Payload cannot be empty"},status=400)
            og_file = TempFiles.objects.create(file=request.FILES.get('file'))
            
            name, path = og_file.file.name.split('/')[1], og_file.file.path
            files = [
                ('file', (name,  open(path, 'rb'), 'application/pdf'))
            ]
            if request.data.get("api")=="prod":
                url = app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_POST_PROD']
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                }
            else:
                url = app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_POST']
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET']
                }
            response = requests.request("POST", url, headers=headers, data=payload, files=files)
            response = response.json()
            try:
                if response.get("responseCode")=="SRS016":
                    return Response(data={"status": True,"message":response},status=200)
                else:
                    return Response(data={"status": False,"message":response},status=400)
            except Exception:
                return Response(data={"status": True,"message":response},status=200)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)

    def get(self,request):
        try:
            if request.GET.get("referenceId") == None:
                return Response(data={"status": False,"message":"Reference ID cannot be empty."},status=400)
           
            ref_id = request.GET.get("referenceId")
            user = request.GET.get("user")
            if user not in app_settings.api_users:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
            payload={}

            if request.GET.get("api")=="prod":
                url = app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET_PROD'] + request.GET.get("referenceId")
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                }
                data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'],"payload":ref_id,"api_name":"Credit Balance Sheet Report - GET"}

            else:
                url = app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'] + request.GET.get("referenceId")
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET']
                    }
                data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'],"payload":ref_id,"api_name":"Credit Balance Sheet Report - GET"}


            response = requests.request("GET", url, headers=headers, data=payload)
            response = json.loads(response.text)
            data['response'] = response
            # data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'],"payload":ref_id,"response":response,"api_name":"Credit Balance Sheet Report - GET"}
            HitSaver(data)
            try:
                if response.get("responseCode")=="SRC001" and response.get("data").get("responseCode")=="SRC001":
                    if request.GET.get("data_type")=="excel":
                        url = response['data']['fileUrl']
                        response = requests.request("GET", url, headers=headers, data=payload)
                        filename = ref_id + ".xlsx"

                        with open(filename, "wb") as file:
                            file.write(response.content)
                        from openpyxl import load_workbook

                        excel_file_path = filename

                        workbook = load_workbook(excel_file_path)

                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            images_to_remove = []
                            for image in sheet._images:
                                images_to_remove.append(image)

                            for image in images_to_remove:
                                sheet._images.remove(image)
                        workbook.save(ref_id + ".xlsx")
                        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        response["Content-Disposition"] = "attachment; filename={}".format(ref_id)
                        with open(ref_id + ".xlsx", "rb") as file:
                            response.write(file.read())
                        os.remove(ref_id + ".xlsx")
                        return response

                    else:
                        # return Response(data={"status": True,"message":response},status=200)
                        report_url=response.get('data').get('jsonUrl')
                        response = requests.request("GET",report_url, headers=headers)
                        return Response(data={"status": True,"message":(json.loads(response.text))},status=200)
                else:
                    return Response(data={"status": False,"message":response},status=400)
                # elif response.get("responseCode")=="EIF009":
                #     return Response(data={"status": False,"message":{"message":"Incorrect File","response_code":"EIF009"}},status=400)
                # elif response.get("responseCode")=="RNP020":
                #     return Response(data={"status": False,"message":{"message":"Request under process","response_code":"RNP020"}},status=400)
                # elif response.get("message")=="Credits not available.":
                #     return Response(data={"status": False,"message":{"message":"Unable to process due to technical issues.","response_code":"400"}},status=400)
            except Exception:
                return Response(data={"status": False,"message":response},status=400)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)

class Gstscore(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        try:
            user = request.GET.get("user")
            if user in app_settings.api_users:
                gst_number = request.GET.get("gst_number")
                if len(gst_number) !=15:
                    return Response(data={"status": False,"message":"Please enter a valid G.S.T number."},status=400)
                
                api = request.GET.get("api")
                if api == 'prod':
                    url = app_settings.EXTERNAL_API['GST_SCORE_ME_PROD']
                    headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD'],
                    'Content-Type': 'application/json'
                    }
                else:
                    url = app_settings.EXTERNAL_API['GST_SCORE_ME']
                    headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET'],
                    'Content-Type': 'application/json'
                    }
                payload = json.dumps({"gstin" : gst_number})

                response = requests.request("POST", url, headers=headers, data=payload)
                response = response.json()
                data={"user":user,"api_endpoint":url,"payload":payload,"response":response,"api_name":"Gstscore"}
                HitSaver(data)
                try:
                    return Response(data={"status": True,"message":(response['data'])},status=200)
                except Exception:
                    return Response(data={"status": False,"message":"API down or No data available for provided G.S.T Number"},status=400)
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class IcwaiByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                membership_id = request.data.get("membership_id")
                if membership_id == None or membership_id =='':
                    return Response(data={"status": False,"message":"Key - membership_id cannot be blank."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                if request.data.get("api")=="prod":
                    url = endpoint_details['ICWAI_URL']
                else:
                    url = endpoint_details['ICWAI_URL_TEST']

                payload = json.dumps({
                "consent": "Y",
                "membership_no": str(membership_id)
                })

                data = KarzaFunction(user, 'icwai-membership', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
        
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class CaMembershipAuthByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                membership_id = request.data.get("membership_id")
                if membership_id == None or membership_id =='':
                    return Response(data={"status": False,"message":"Key - membership_id cannot be blank."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                
                if request.data.get("api")=="prod":
                    url = endpoint_details['CA_MEMBERSHIP']
                else:
                    url = endpoint_details['CA_MEMBERSHIP_TEST']

                payload = json.dumps({
                "membership_no": str(membership_id),
                "consent": "y"
                })
                data = KarzaFunction(user,'ca-membership', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)         
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class ShopAndEstablishmentByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                area_code = request.data.get("area_code")
                registration_no = request.data.get("registration_no")
                if area_code == None or area_code =='':
                    return Response(data={"status": False,"message":"Key - area_code cannot be blank."},status=400)
                elif registration_no == None or registration_no =='':
                    return Response(data={"status": False,"message":"Key - registration_no cannot be blank."},status=400)

                endpoint_details = app_settings.EXTERNAL_API

                if request.data.get("api")=="prod":
                    url = endpoint_details['SHOP_AND_EST']
                else:
                    url = endpoint_details['SHOP_AND_EST_TEST']

                payload = json.dumps({
                "regNo": registration_no,
                "pdfRequired": False,
                "areaCode": str(area_code),
                "consent": "Y"
                })

                data = KarzaFunction(user,'shop-establishment', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)      
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class FssaiAuthByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                reg_no = request.data.get("reg_no")
                if reg_no == None or reg_no =='':
                    return Response(data={"status": False,"message":"Key - membership_id cannot be blank."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                url = endpoint_details['FSSAI_TEST']
                if request.data.get("api")=="prod":
                    url = endpoint_details['FSSAI']
                else:
                    url = endpoint_details['FSSAI_TEST']

                payload = json.dumps({
                "reg_no": str(reg_no),
                "consent": "y"
                })

                data = KarzaFunction(user, 'fssai-membership',url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class BankVerificationByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                accountNumber = request.data.get("account_no")
                ifsc = request.data.get("ifsc")
                if accountNumber == None or accountNumber =='' or ifsc == None or ifsc =='' :
                    return Response(data={"status": False,"message":"Key - ifsc and accountNumbern cannot be blank."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                
                if request.data.get("api")=="prod":
                    url = endpoint_details['BANK_VERIFICATION']
                else:
                    url = endpoint_details["BANK_VERIFICATION_TEST"]

                payload = json.dumps({
                "consent": "Y",
                "ifsc": ifsc,
                "accountNumber": accountNumber
                })
                data = KarzaFunction(user, 'bank-verification', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class UdyamRegCheckByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                registration_no = request.data.get("registration_no")
                if registration_no == None or registration_no =='':
                    return Response(data={"status": False,"message":"Key - registration_no cannot be blank."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                
                if request.data.get("api")=="prod":
                    url = endpoint_details['UDHYAM_AUTH']
                else:
                    url = endpoint_details["UDHYAM_AUTH_TEST"]

                payload = json.dumps({
                "consent": "Y",
                "udyamRegistrationNo": registration_no,
                "isPDFRequired": "N",
                "getEnterpriseDetails": "Y"
                })
                data = KarzaFunction(user, 'udhyam-verification', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class LEIByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                lei_no = request.data.get("lei_no")
                if lei_no == None or lei_no =='':
                    return Response(data={"status": False,"message":"Key - lei_no cannot be blank."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                
                if request.data.get("api")=="prod":
                    url = endpoint_details['LEI']
                else:
                    url = endpoint_details["LEI_TEST"]

                payload = json.dumps({
                "leiNo": lei_no,
                "consent": "Y"
                })
                data = KarzaFunction(user, 'lei', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class ITRAuthByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                acknowledgement_no = request.data.get("acknowledgement_no")
                if acknowledgement_no == None or acknowledgement_no =='' or len(acknowledgement_no)!=15:
                    return Response(data={"status": False,"message":"Key - acknowledgement_no should be exact 15 digits in length."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                
                if request.data.get("api")=="prod":
                    url = endpoint_details['ITR']
                else:
                    url = endpoint_details["ITR_TEST"]

                payload = json.dumps({
                "ack": acknowledgement_no,
                "consent": "Y"
                })
                data = KarzaFunction(user, 'ITR-Verification', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class PassportByk(APIView):
    # permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                file_no = request.data.get("file_no")
                dob = request.data.get("dob")
                if file_no == None or file_no =='' or dob == None or dob =='':
                    return Response(data={"status": False,"message":"Key - file_no and dob both are required."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                url = endpoint_details["KARZAAPI_PASSPORT"]

                payload = json.dumps({
                "fileNo": file_no,
                "dob": dob,
                "consent": "y"
                })
                data = KarzaFunction(user, 'Passport', url, payload, endpoint_details)
                result_mod = {}
                if data['status']==True:
                    try:
                        fname= data['message']['result']['name']['nameFromPassport']
                    except Exception:
                        fname= ""

                    try:
                        lname = data['message']['result']['name']['surnameFromPassport']
                    except Exception:
                        lname = ''
                    result_mod['Name'] = fname + " " + lname
                    result_mod['Passport No.'] = data['message']['result']['passportNumber']['passportNumberFromSource']
                    result_mod['Date of issue'] = data['message']['result']['dateOfIssue']['dispatchedOnFromSource']
                    result_mod['File No.'] = request.data.get("file_no")
                    result_mod['DOB'] = request.data.get("dob")
                return Response(data = {'status' : data['status'], 'message' : result_mod}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        
class EpfByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                cin_id = request.data.get("cin_id")
                from_date = request.data.get("from_date")
                to_date = request.data.get("to_date")
                if None in [cin_id, from_date, to_date] or '' in [cin_id, from_date, to_date]:
                    return Response(data={"status": False,"message":"Key - cin_id, from_date, to_date are compulsory."},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                
                if request.data.get("api")=="prod":
                    url = endpoint_details['EPF_COMP']
                else:
                    url = endpoint_details["EPF_COMP_TEST"]

                payload = json.dumps({
                "user":user,
                "cinKid": cin_id,
                "fromDate": from_date,
                "toDate": to_date
                })
                data = KarzaFunction(user, 'EPF', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']['result']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        

def CustomOrderDetails(request):
    try:
        
        from hv_whatsapp_api.models import order
        data=list(order.objects.filter().values('order_id','mobile_no','name','customer_info__service_type_id',"price",'transaction_id','payment_recieved_date','auto_or_manual','final_status','report_sent_time'))

        from django.http import HttpResponse
        import pandas as pd
        result_mod = pd.DataFrame(data)
        result_mod.rename(columns={"customer_info__service_type_id":"Service_type"},inplace=True)
        package_name={21:'Verify Nanny',22:'Verify Driver',23:'Verify domestichelp/Security guard', 24:'Verify Anyone', 25:'Know the identity',26:'Profile your date', 27:'Get cyber identity verified',28:'Get verified badge'}
        
        result_mod['Service_type'].replace(package_name, inplace=True)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
        result_mod.to_excel(response, index=False)
        return response
    except Exception:
        return Response(data={"status": False,"message":"API down"},status=400)
    
class UANByK(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        # user = request.data.get("user","Unknown")
        # if user in app_settings.api_users:
        phone = request.data.get("phone")
        if phone == "" or phone == None:
            phone = 0
        uan = request.data.get("uan")
        if uan == "" or uan == None:
            uan = 0
        name = request.data.get("name")
        if name == "" or name == None:
            name = 0
        user = request.data.get("user")
        result = UanHandler(phone, uan, name)

        if len(result)==1: #Uan provided
            try:
                result_mod = result[0]['result']['uan'][0]['employer']
                final_data={}
                for i in range(0,len(result_mod)):
                    final_data['Name - '+ str(i+1)] = result_mod[i]['name']
                    final_data['Member ID - '+ str(i+1)] = result_mod[i]['memberId']
                    final_data['Date of exit - '+ str(i+1)] = result_mod[i]['dateOfExit']
                    final_data['Date of joining - '+ str(i+1)] = result_mod[i]['dateOfJoining']
                final_data['UAN'] = uan
                data={"user":user,"api_endpoint":"https://api.karza.in/v2/employment-verification & https://api.karza.in/v2/uan-lookup","payload":(str(phone) + str(name) + str(uan)),"response":(str(result) + "======///======" + str(final_data)) ,"api_name":"UAN API"}

                HitSaver(data)
                return Response(data={"status": True,"message":final_data},status=200)
                #list received here
            except Exception:
                result_mod = {"Message" : "No data found."}
                data={"user":user,"api_endpoint":"https://api.karza.in/v2/employment-verification & https://api.karza.in/v2/uan-lookup","payload":(str(phone) + str(name) + str(uan)),"response":(str(result) + "======///======" + str(result_mod)) ,"api_name":"UAN API"}
                HitSaver(data)
                return Response(data={"status": False,"message":result_mod},status=400)

        elif len(result)==2:
            try:
                uan_nos = result[0]
                uan_data = result[1]
                result_mod={}
                just_data=[]
                for i in range(0,len(uan_data)):
                    just_data.append(uan_data[i]['result']['uan'][0]['employer'])
                    
                        
                # print(just_data )   
                new_data=[]
                for i in range(0, len(just_data)):
                    x=just_data[i]
                    for j in range(0,len(x)):
                        new_data.append(x[j])

                uan_co=[]
                for i in just_data:
                    uan_co.append(len(i))
                    
                result1 = []

                # Calculate the running sum
                running_sum = 0
                for num in uan_co:
                    running_sum += num
                    result1.append(running_sum)

                for i in range(0,len(new_data)):
                        result_mod['Name - '+ str(i+1)] = new_data[i]['name']
                        result_mod['Member ID - '+ str(i+1)] = new_data[i]['memberId']
                        result_mod['Date of exit - '+ str(i+1)] = new_data[i]['dateOfExit']
                        result_mod['Date of joining - '+ str(i+1)] = new_data[i]['dateOfJoining']
                        result_mod['------'+str(i+1)] = '------'+str(i+1)
                        for k in range(0,len(uan_nos)):
                            if i+1 == result1[k]:
                                result_mod['UAN for above ' + "company/companies is" + (" ")*k] = uan_nos[k]
                
                data={"user":user,"api_endpoint":"https://api.karza.in/v2/employment-verification & https://api.karza.in/v2/uan-lookup","payload":(str(phone)+"///" + str(name) +"///"+ str(uan)),"response":(str(result) + "======///======" + str(result_mod)) ,"api_name":"UAN API"}
                HitSaver(data)
                return Response(data={"status": True,"message":result_mod},status=200)
                    
            except Exception:
                result_mod = {"Message" : "No data found, Please retry later"}
                data={"user":user,"api_endpoint":"https://api.karza.in/v2/employment-verification & https://api.karza.in/v2/uan-lookup","payload":(str(phone)+"///" + str(name) +"///"+ str(uan)),"response":(str(result) + "======///======" + str(result_mod)) ,"api_name":"UAN API"}

                HitSaver(data)
                return Response(data={"status": True,"message":result_mod},status=200)

class PassportOCR(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        from ocr_apis.utils import PassportIDProcessor
        try:
            data = detect_text(request.data.get('file').read())
            if request.data.get('id') =='front':
                result = PassportIDProcessor.get_passport_front_result(self,data)
                
            else:
                result = PassportIDProcessor.get_passport_back_result(self,data)
            
            return Response(data={"status": True,"message":result},status=200)
        except Exception:
            return Response(data={"status": False,"message":"Failed to read the image"},status=400)
        
class gstreport(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user") 
            pan = request.data.get("pan")
            if user in app_settings.api_users:
                if len(pan)==10:

                    url = app_settings.EXTERNAL_API['GST_REPORT_SCOREME']
                    payload = json.dumps({
                        "pan": pan
                        })
                    headers = {
                        'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                        'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD'],
                        'Content-Type': 'application/json'
                        }
                    
                    response = requests.request("POST", url, headers=headers, data=payload)
                    response = response.json()
                    data={"user":user,"api_endpoint":url,"payload":payload,"response":response ,"api_name":"GST-Report: Submit PAN"}
                    HitSaver(data)
                    if type(response.get('data')) != None: 
                        return Response(data={"status": True,"message":response},status=200)
                    else:
                        return Response(data={"status": False,"message":'API issue/incorrect data provided. Please retry.'},status=400)
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":'API down'},status=400)
        
    def get(self, request):
        try:
            user = request.GET.get("user") 
            ref_id = request.GET.get("ref_id") 
            if user in app_settings.api_users and ref_id!='':
                url = app_settings.EXTERNAL_API['GST_REPORT_DOWNLOAD_SCOREME']+"?referenceId=" + str(ref_id)
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                    }
                payload ={}
                response = requests.request("GET", url, headers=headers, data=payload)
                response = response.json()
                data={"user":user,"api_endpoint":url,"payload":ref_id,"response":response ,"api_name":"GST-Report: Download report"}
                HitSaver(data)
                if response.get('data').get('jsonUrl')!=None:
                    if request.GET.get("data_type")=="excel":
                        url =response['data']['excelUrl'] 
                        response = requests.request("GET", url, headers=headers, data=payload)
                        filename = ref_id + ".xlsx"

                        with open(filename, "wb") as file:
                            file.write(response.content)
                        from openpyxl import load_workbook

                        excel_file_path = filename

                        workbook = load_workbook(excel_file_path)

                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            images_to_remove = []
                            for image in sheet._images:
                                images_to_remove.append(image)

                            for image in images_to_remove:
                                sheet._images.remove(image)
                        workbook.save(ref_id + ".xlsx")
                        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        response["Content-Disposition"] = "attachment; filename={}".format(ref_id)
                        with open(ref_id + ".xlsx", "rb") as file:
                            response.write(file.read())
                        os.remove(ref_id + ".xlsx")
                        return response

                    else:
                        report_url=response.get('data').get('jsonUrl')
                        response = requests.request("GET",report_url, headers=headers)
                        return Response(data={"status": True,"message":((json.loads(response.text)))},status=200)
        except Exception:
            return Response(data={"status": False,"message":"Response not generated yet/Please recheck the reference ID you have provided."},status=400)

class McaByk(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                cin_id = request.data.get("cin_id")
                if len(cin_id)<1:
                    return Response(data={"status": False,"message":"CIN cannot be empty"},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API
                
                
                url = endpoint_details['MCA_KARZA']
                

                payload = json.dumps({
                "consent": "Y",
                "cin": cin_id
                })
                data = KarzaFunction(user, 'EPF', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']['result']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)



#------------------------------------------------------------------------------------------------------------------------------------------------------------

class CompanyReportRequest(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self,request):
        try:
            if request.data.get('user') not in app_settings.api_users:
                return Response(data={"status": True,"message":"User not authorized"},status=400)
            
            payload = json.dumps({
                "cin_llpin": request.data.get('cin')
            })

            """ 
            url = "https://quality-da.scoreme.in/mca/external/companyReportRequest"
            headers = {
            'clientId': '800d9a0bc1d65dc51c607d072706df03',
            'clientSecret': '04b843cb9d4f9eb2af61ea8ae95922b9f20d51b133da7cd4dfd64db2ea678b15',
            'Content-Type': 'application/json'
            } """

            if request.data.get('cin')=='' or request.data.get('cin')==None:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
            
            if request.data.get("api")=="prod": #SCORE_ME_COMPNY_REPORT_REQUEST_PROD
                # url = "https://da.scoreme.in/mca/external/companyReportRequest"
                url = app_settings.EXTERNAL_API['SCORE_ME_COMPNY_REPORT_REQUEST_PROD']
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD'],
                    'Content-Type': 'application/json'
                    }
            else:
                url = "https://quality-da.scoreme.in/mca/external/companyReportRequest"
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET'],
                    'Content-Type': 'application/json'
                    }
            
            response = requests.request("POST", url, headers=headers, data=payload)
            response= response.json()
            try:
                if response.get("responseCode")=="SRS016":
                    return Response(data={"status": True,"message":response},status=200)
                else:
                    return Response(data={"status": False,"message":response},status=400)
            except Exception:
                return Response(data={"status": True,"message":response},status=200)
        except:
            return Response(data={"status": False,"message":"API down"},status=400)

    def get(self,request):
        try:
            if request.GET.get("referenceId") == None:
                return Response(data={"status": False,"message":"Reference ID cannot be empty."},status=400)
            
            ref_id = request.GET.get("referenceId")
            user = request.GET.get("user")
            if user not in app_settings.api_users:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
            payload={}

            if request.GET.get("api")=="prod":
                url = app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET_PROD'] + request.GET.get("referenceId")
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                }
                data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'],"payload":ref_id,"api_name":"Credit Balance Sheet Report - GET"}

            else:
                url = app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'] + request.GET.get("referenceId")
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET']
                    }
                data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'],"payload":ref_id,"api_name":"Credit Balance Sheet Report - GET"}


            response = requests.request("GET", url, headers=headers, data=payload)
            response = json.loads(response.text)
            data['response'] = response
            # data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'],"payload":ref_id,"response":response,"api_name":"Credit Balance Sheet Report - GET"}
            HitSaver(data)
            try:
                if response.get("responseCode")=="SRC001" and response.get("data").get("responseCode")=="SRC001":
                    if request.GET.get("data_type")=="excel":
                        url = response['data']['excelUrl']
                        response = requests.request("GET", url, headers=headers, data=payload)
                        filename = ref_id + ".xlsx"

                        with open(filename, "wb") as file:
                            file.write(response.content)
                        from openpyxl import load_workbook

                        excel_file_path = filename

                        workbook = load_workbook(excel_file_path)

                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            images_to_remove = []
                            for image in sheet._images:
                                images_to_remove.append(image)

                            for image in images_to_remove:
                                sheet._images.remove(image)
                        workbook.save(ref_id + ".xlsx")
                        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        response["Content-Disposition"] = "attachment; filename={}".format(ref_id)
                        with open(ref_id + ".xlsx", "rb") as file:
                            response.write(file.read())
                        os.remove(ref_id + ".xlsx")
                        return response

                    else:
                        # return Response(data={"status": True,"message":response},status=200)
                        report_url=response.get('data').get('jsonUrl')
                        response = requests.request("GET",report_url, headers=headers)
                        return Response(data={"status": True,"message":(json.loads(response.text))},status=200)
                else:
                    return Response(data={"status": False,"message":response},status=400)
                # elif response.get("responseCode")=="EIF009":
                #     return Response(data={"status": False,"message":{"message":"Incorrect File","response_code":"EIF009"}},status=400)
                # elif response.get("responseCode")=="RNP020":
                #     return Response(data={"status": False,"message":{"message":"Request under process","response_code":"RNP020"}},status=400)
                # elif response.get("message")=="Credits not available.":
                #     return Response(data={"status": False,"message":{"message":"Unable to process due to technical issues.","response_code":"400"}},status=400)
            except Exception:
                return Response(data={"status": False,"message":response},status=400)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)

class GSTReportRequest(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self,request):
        try:
            if request.data.get('user') not in app_settings.api_users:
                return Response(data={"status": True,"message":"User not authorized"},status=400)
            
            payload = json.dumps({
                "pan": request.data.get('pan_no')
            })

            """ 
            url = "https://sm-gst.scoreme.in/gst/external/search/gstinFilingStatusByPan"
            headers = {
            'clientId': '800d9a0bc1d65dc51c607d072706df03',
            'clientSecret': '04b843cb9d4f9eb2af61ea8ae95922b9f20d51b133da7cd4dfd64db2ea678b15',
            'Content-Type': 'application/json'
            } """

            if request.data.get('pan_no')=='' or request.data.get('pan_no')==None:
                return Response(data={"status": False,"message":"Data not provided properly!"},status=400)
            
            if request.data.get("api")=="prod":#SCORE_ME_GST_IN_FILING_STATUS_BY_PAN_PROD
                # url = "https://sm-gst.scoreme.in/gst/external/search/gstinFilingStatusByPan"
                url = app_settings.EXTERNAL_API['SCORE_ME_GST_IN_FILING_STATUS_BY_PAN_PROD']
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD'],
                    'Content-Type': 'application/json'
                    }
            else:
                url = "https://sm-gst-sandbox.scoreme.in/gst/external/search/gstinFilingStatusByPan"
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET'],
                    'Content-Type': 'application/json'
                    }
            
            response = requests.request("POST", url, headers=headers, data=payload)
            response= response.json()
            try:
                if response.get("responseCode")=="SRS016":
                    return Response(data={"status": True,"message":response},status=200)
                else:
                    return Response(data={"status": False,"message":response},status=400)
            except Exception:
                return Response(data={"status": True,"message":response},status=200)
        except:
            return Response(data={"status": False,"message":"API down"},status=400)

    def get(self,request):
        try:
            if request.GET.get("referenceId") == None:
                return Response(data={"status": False,"message":"Reference ID cannot be empty."},status=400)
            
            ref_id = request.GET.get("referenceId")
            user = request.GET.get("user")
            if user not in app_settings.api_users:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
            payload={}

            if request.GET.get("api")=="prod":
                url = app_settings.EXTERNAL_API['GST_REPORT_DOWNLOAD_SCOREME'] + request.GET.get("referenceId")
                headers = {
                'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID_PROD'],
                'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET_PROD']
                }
                data={"user":user,"api_endpoint":app_settings.EXTERNAL_API['GST_REPORT_DOWNLOAD_SCOREME'],"payload":ref_id,"api_name":"GST By Pan Report - GET"}

            else:
                url = app_settings.EXTERNAL_API['CREDIT_BALANCE_SHEET_GET'] + request.GET.get("referenceId")
                url = "https://sm-gst-sandbox.scoreme.in/gst/external/getgstreport"+"?referenceId=" + str(ref_id)
                headers = {
                    'clientId': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_ID'],
                    'clientSecret': app_settings.EXTERNAL_API['SCORE_ME_CLIENT_SECRET']
                    }
                data={"user":user,"api_endpoint":"https://sm-gst-sandbox.scoreme.in/gst/external/getgstreport","payload":ref_id,"api_name":"GST By Pan Report - GET"}


            response = requests.request("GET", url, headers=headers, data=payload)
            response = json.loads(response.text)
            data['response'] = response
            HitSaver(data)
            try:
                if response.get("responseCode")=="SRC001" or response.get("data").get("responseCode")=="SRC001":
                    if request.GET.get("data_type")=="excel":
                        url = response['data']['excelUrl']
                        response = requests.request("GET", url, headers=headers, data=payload)
                        filename = ref_id + ".xlsx"

                        with open(filename, "wb") as file:
                            file.write(response.content)
                        from openpyxl import load_workbook

                        excel_file_path = filename

                        workbook = load_workbook(excel_file_path)

                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            images_to_remove = []
                            for image in sheet._images:
                                images_to_remove.append(image)

                            for image in images_to_remove:
                                sheet._images.remove(image)
                        workbook.save(ref_id + ".xlsx")
                        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        response["Content-Disposition"] = "attachment; filename={}".format(ref_id)
                        with open(ref_id + ".xlsx", "rb") as file:
                            response.write(file.read())
                        os.remove(ref_id + ".xlsx")
                        return response

                    else:
                        # return Response(data={"status": True,"message":response},status=200)
                        report_url=response.get('data').get('jsonUrl')
                        response = requests.request("GET",report_url, headers=headers)
                        return Response(data={"status": True,"message":(json.loads(response.text))},status=200)
                else:
                    return Response(data={"status": False,"message":response},status=400)
            except Exception:
                return Response(data={"status": False,"message":response},status=400)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)

class MCASignatories(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        try:
            user = request.data.get("user","Unknown")
            if user in app_settings.api_users:
                cin_id = request.data.get("cin_id")
                if len(cin_id)<1:
                    return Response(data={"status": False,"message":"CIN cannot be empty"},status=400)
                
                endpoint_details = app_settings.EXTERNAL_API # https://testapi.karza.in/v2/mca-signatories
                url = endpoint_details['MCA_SIGNATORIES']
                
                payload = json.dumps({
                "consent": "Y",
                "cin": cin_id
                })
                data = KarzaFunction(user, 'EPF', url, payload, endpoint_details)
                return Response(data = {'status' : data['status'], 'message' : data['message']['result']}, status = data['status_no'])
            else:
                return Response(data={"status": False,"message":"User not registered for API usage"},status=401)
        except Exception:
            return Response(data={"status": False,"message":"API down"},status=400)
        


class PassportOcrWithMrzCode(APIView):   # Passport ocr with mrz code
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        if not request.data.get('passport_front') and not request.data.get('passport_back') :
            return Response(data={"status": False,"message":"Passport front and Back is mindetry."},status=400)
        front_result = {}
        back_result = {}
        result = {}
        try:
            front_ocr = detect_text(request.data.get('passport_front').read())
            front_result = get_indian_passport_front_result(front_ocr)
            back_ocr = detect_text(request.data.get('passport_back').read())
            back_result = get_indian_passport_back_result(back_ocr)
              
            result = margedict(front_result,back_result)
           
            data = {
                'client': request.data.get('user'),
                'passport_ocr_json': result
            }
            return Response(data={"status": True,"message":data},status=200)
        except Exception as ex:
            return Response(data={"status": False,"message": "Someting went wrong, {}.".format(str(ex))},status=400)
        
class PanCardOcr(APIView):   # Pan card ocr code
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        if not request.data.get('pancard_img'):
            return Response(data={"status": False,"message":"Pan card is mindetry."},status=400)
        result = {}
        try:
            front_ocr = detect_text(request.data.get('pancard_img').read())
            result = get_pan_card_ocr_json_result(front_ocr)
            
           
            data = {
                'client': request.data.get('user'),
                'pancard_ocr_json': result
            }
            return Response(data={"status": True,"message":data},status=200)
        except Exception as ex:
            return Response(data={"status": False,"message": "Someting went wrong, {}.".format(str(ex))},status=400)
        
class VoterIdOcr(APIView):   # Voter Id ocr with code
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        if not request.data.get('voter_front') and not request.data.get('voter_back') :
            return Response(data={"status": False,"message":"Voter Id front and Back is mindetry."},status=400)
        result = {}
        try:
            front_ocr = detect_text(request.data.get('voter_front').read())
            front_result = get_voter_id_front_json(front_ocr)
            if request.data.get('voter_back') and request.data.get('voter_back') != '':
                back_ocr = detect_text(request.data.get('voter_back').read())
                back_result = get_voter_id_back_json(back_ocr)
            else:
                back_result = {
                    'voter_id_num':'', 
                    'address':'', 
                    'issue_date':''
                    }    
            result = margedict(front_result,back_result)
           
            data = {
                'client': request.data.get('user'),
                'voter_id_ocr_json': result            }
            return Response(data={"status": True,"message":data},status=200)
        except Exception as ex:
            return Response(data={"status": False,"message": "Someting went wrong, {}.".format(str(ex))},status=400)
        
class DLOcrCode(APIView):   # Dl ocr with code
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        from external_apis.dl_regex import get_dl_front_json, get_dl_back_json
        if not request.data.get('dl_front') :
            return Response(data={"status": False,"message":"Driving License  front is mindetry."},status=400)
        result = {}
        try:
            front_ocr = detect_text(request.data.get('dl_front').read())
            front_result = get_dl_front_json(front_ocr)
            if request.data.get('dl_back') and request.data.get('dl_back') != '':
                back_ocr = detect_text(request.data.get('dl_back').read())
                back_result = get_dl_back_json(back_ocr)
            else:
                 back_result = {
                    'dl_num':'', 
                    'address':''
                    }   
            
            result = margedict(front_result,back_result)
           
            data = {
                'client': request.data.get('user'),
                'dl_ocr_json': result            }
            return Response(data={"status": True,"message":data},status=200)
        except Exception as ex:
            return Response(data={"status": False,"message": "Someting went wrong, {}.".format(str(ex))},status=400)                        


class AadhaarOcrCode(APIView):   # Aadhaar ocr with code
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request):
        
        if not request.data.get('aadhaar_front') and not request.data.get('aadhaar_back') :
            return Response(data={"status": False,"message":"Aadhaar card  front and Back is mindetry."},status=400)
        result = {}
        try:
            front_ocr = detect_text(request.data.get('aadhaar_front').read())
            front_result = get_aadhaar_front_json(front_ocr)
            if request.data.get('aadhaar_back') and request.data.get('aadhaar_back') != '':
                back_ocr = detect_text(request.data.get('aadhaar_back').read())
                back_result = get_aadhaar_back_json(back_ocr)
            else:
                 back_result = {
                    'aadhaar_num':'', 
                    'address':''
                    }   
            
            result = margedict(front_result,back_result)
           
            data = {
                'client': request.data.get('user'),
                'aadhaar_ocr_json': result            }
            return Response(data={"status": True,"message":data},status=200)
        except Exception as ex:
            return Response(data={"status": False,"message": "Someting went wrong, {}.".format(str(ex))},status=400)  

class DocumentFroud(APIView):
    permission_classes = [permissions.AllowAny]
    def post(self, request):
        try:
            print("Doc Froud...")
            username = "0oaatrbjbeHqo7Dk0417"
            password = "bZoUL3eMGUxkGXofizoJIlbXl9AIdkXnQ5IF0w7XyT8Hat31TSkmumYZAPIcxfxq"
            encoded_credentials = encode_credentials(username, password)

            url = "https://eu.id.resistant.ai/oauth2/aus17c52xbW6c0yA9417/v1/token"

            payload = 'grant_type=client_credentials&scope=submissions.read%20submissions.write'
            headers = {
            'Content-Type': 'application/x-www-form-urlencoded',
            'Accept': 'application/json',
            'Authorization': f'Basic {encoded_credentials}'  # Use the encoded credentials
            }

            response = requests.request("POST", url, headers=headers, data=payload)

            if response.status_code == 200:
                print("User authenticated successfully for access token.")
                url2 = "https://api.documents.testing.resistant.ai/v2/submission"
                headers2 = {
                    'Content-Type': 'application/json',
                    'Authorization': 'Bearer'+' '+ json.loads(response.text)['access_token'] 
                    }
                payload = json.dumps({
                    "query_id": "50",
                    "pipeline_configuration": "FRAUD_ONLY",
                    "enable_decision": False,
                    "enable_submission_characteristics": False})
                
                response2 = requests.request("POST", url2, headers=headers2, data=payload)
                if response2.status_code == 200:
                    print("Upload url and submission id generated successfully.")
                    data = json.loads(response2.text)
                    return Response(data={"status": True,"message":data},status=200)
        except Exception as ex:
            return Response(data={"status": False,"message": "Someting went wrong, {}.".format(str(ex))},status=400)



class DocumentFroudResult(APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        print("Doc Froud...")
        logging.debug("Doc Froud...")
        if not request.FILES.get('file'):
            return Response(data={"status": False, "message": "Document file is mandatory."}, status=400)
        
        try:
            url = "https://checkapi.helloverify.com/external-apis/document_froud"
            headers = {}
            payload = {}
            generate_credential = requests.request("POST", url, headers=headers, data=payload)
            if generate_credential.status_code == 200:
                actual_res = (json.loads(generate_credential.text))
                request.session['froud_submission_id'] = actual_res['message']['submission_id']
                request.session['froud_upload_url'] = actual_res['message']['upload_url']
        except Exception as ex:
            print(str(ex))
            logging.debug("Generate credential status..{}.".format(str(generate_credential.text)))
            return Response(data={"status": False, "message": "Something went wrong, {}.".format(str(ex))}, status=400)
                
        print("document_froud_analysis_result endpoint...")
        if request.FILES.get('file'):
            try:
                url = request.session['froud_upload_url']
                payload = request.FILES['file']
                headers = {
                    'Content-Type': 'application/octet-stream'
                }

                upload_response = requests.request("PUT", url, headers=headers, data=payload)
                logging.debug("Document uploaded status..{}.".format(str(upload_response.text)))
                if upload_response.status_code == 200:
                    print("Document uploaded successfully.")
                    # Use the function to create encoded credentials
                    username = "0oaatrbjbeHqo7Dk0417"
                    password = "bZoUL3eMGUxkGXofizoJIlbXl9AIdkXnQ5IF0w7XyT8Hat31TSkmumYZAPIcxfxq"
                    encoded_credentials = encode_credentials(username, password)
                    
                    url = "https://eu.id.resistant.ai/oauth2/aus17c52xbW6c0yA9417/v1/token"
                    payload = 'grant_type=client_credentials&scope=submissions.read%20submissions.write'
                    headers = {
                        'Content-Type': 'application/x-www-form-urlencoded',
                        'Accept': 'application/json',
                        'Authorization': f'Basic {encoded_credentials}'  # Use the encoded credentials
                    }

                    authentication_response = requests.request("POST", url, headers=headers, data=payload)
                    if authentication_response.status_code == 200:
                        print("User authenticated successfully.")
                        submission_id = request.session['froud_submission_id']
                        DocumentTampering.objects.create(submission_id=submission_id,initial_doc=request.FILES['file'])
                        analysis_result = self.check_fraud_analysis(submission_id, json.loads(authentication_response.text)['access_token'])
                        if analysis_result.status_code== 200:
                            return self.draw_multiple_bounding_boxes(submission_id,analysis_result)
                        return analysis_result
                    return Response(data={"status": False, "message": "We got error into authentication endpoint, {}.".format(str(authentication_response.text))}, status=400)    
                return Response(data={"status": False, "message": "We got error into Upload document endpoint , {}.".format(str(upload_response.text))}, status=400)        
            except Exception as ex:
                print(traceback.print_exc())
                logging.warning("<----------"+str(datetime.now())+"---------->")
                logging.exception((inspect.currentframe().f_code.co_name).upper())
                return Response(data={"status": False, "message": "Got Error in Document Tampering Document Upload...  , {}.".format(str(ex))}, status=400)
        return Response(data={"status": False, "message": "upload_url, submission_id and Document file are mandatory."}, status=400)

    def check_fraud_analysis(self, submission_id, access_token, retries=6, delay=10):
        import time
        """
        Function to recursively check the fraud analysis result.
        """
        try:
            url = f"https://api.documents.testing.resistant.ai/v2/submission/{submission_id}/fraud?with_metadata=True"
            headers = {
                'Authorization': f'Bearer {access_token}'
            }
            time.sleep(delay)  # wait before making the next request
            final_response = requests.request("GET", url, headers=headers)
            logging.debug("Analysis Result status..{}.".format(str(final_response.text)))
            if final_response.status_code == 200:
                data = json.loads(final_response.text)
                return Response(data={"status": True, "message": data}, status=200)
            elif retries > 0:
                print(f"Retrying... Attempts left: {retries}")
                return self.check_fraud_analysis(submission_id, access_token, retries=retries-1, delay=delay)
            else:
                return Response(data={"status": False, "message": "Maximum retry attempts reached. {}.".format(str(final_response.text))}, status=400)
        except Exception as ex:
            print(traceback.print_exc())
            logging.warning("<----------"+str(datetime.now())+"---------->")
            logging.exception((inspect.currentframe().f_code.co_name).upper())
            return Response(data={"status": False, "message": "Got Error in Document Tampering Fraud Analysis Result...  , {}.".format(str(ex))}, status=400)

    def draw_multiple_bounding_boxes(self, submission_id, analysis_response):
        import fitz  # PyMuPDF for PDFs
        from PIL import Image, ImageDraw
        from io import BytesIO
        from django.core.files.base import ContentFile
        import tempfile
        import os

        try:
            # Fetch the document instance using submission_id
            document_instance = DocumentTampering.objects.filter(submission_id=submission_id).last()
            if not document_instance:
                raise ValueError("Document not found for the given submission_id.")

            file_name = document_instance.initial_doc.name.lower()
            file_content = document_instance.initial_doc.read()

            # Save the file content to a temporary file
            with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                temp_file.write(file_content)
                temp_file.flush()
                temp_file_path = temp_file.name

            analysis_result = analysis_response.data.get("message", {}).get("indicators", [])

            # Function to process bounding box data
            def draw_bounding_boxes(coords_list, draw_func):
                for coords in coords_list:
                    # Handle if coords is a list of bounding boxes
                    if isinstance(coords, list):
                        for bbox in coords:
                            if not isinstance(bbox, dict):  # Make sure bbox is a dict
                                continue
                            x, y = bbox.get('x', 0), bbox.get('y', 0)
                            width, height = bbox.get('width', 0), bbox.get('height', 0)
                            draw_func(x, y, width, height)
                    
                    # Handle if coords is a dictionary
                    elif isinstance(coords, dict):
                        bbox = coords.get("bbox", {})
                        
                        # Check if bbox is also a list (some cases have list inside)
                        if isinstance(bbox, list):
                            for box in bbox:
                                if isinstance(box, dict):
                                    x, y = box.get('x', 0), box.get('y', 0)
                                    width, height = box.get('width', 0), box.get('height', 0)
                                    draw_func(x, y, width, height)
                        elif isinstance(bbox, dict):
                            x, y = bbox.get('x', 0), bbox.get('y', 0)
                            width, height = bbox.get('width', 0), bbox.get('height', 0)
                            draw_func(x, y, width, height)
                        else:
                            print(f"Invalid bbox format: {bbox}, skipping.")
                    
                    else:
                        print(f"Unsupported coords format: {coords}, skipping.")

            # Processing for PDF files
            if file_name.endswith(".pdf"):
                pdf_doc = fitz.open(temp_file_path)
                for page_num in range(pdf_doc.page_count):
                    page = pdf_doc[page_num]
                    page_height = page.rect.height

                    # Drawing function for PDF
                    def draw_pdf_box(x, y, width, height):
                        y_from_bottom = page_height - (y + height)
                        rect = fitz.Rect(x, y_from_bottom, x + width, y_from_bottom + height)
                        page.draw_rect(rect, color=(1, 0, 0), width=2)

                    # Draw bounding boxes on each page
                    for ele in analysis_result:
                        metadata = ele.get("metadata")
                        if metadata and metadata.get("data"):
                            draw_bounding_boxes(metadata["data"], draw_pdf_box)

                # Save the modified PDF and update the document instance
                pdf_output_path = f"tampered_{document_instance.submission_id}.pdf"
                pdf_doc.save(pdf_output_path)
                pdf_doc.close()

                with open(pdf_output_path, "rb") as f:
                    document_instance.tampered_doc.save(f"tampered_{document_instance.submission_id}.pdf", ContentFile(f.read()))
                    document_instance.save()

                os.remove(pdf_output_path)

            # Processing for image files
            elif file_name.endswith((".png", ".jpg", ".jpeg")):
                image = Image.open(temp_file_path)
                image_height = image.height
                draw = ImageDraw.Draw(image)

                # Drawing function for images
                def draw_image_box(x, y, width, height):
                    y_from_bottom = image_height - (y + height)
                    draw.rectangle([x, y_from_bottom, x + width, y_from_bottom + height], outline="red", width=2)

                # Draw bounding boxes on the image
                for ele in analysis_result:
                    metadata = ele.get("metadata")
                    if metadata and metadata.get("data"):
                        draw_bounding_boxes(metadata["data"], draw_image_box)

                # Save the modified image and update the document instance
                img_io = BytesIO()
                image_format = image.format or 'JPEG'  # Default to JPEG if format is not available
                image.save(img_io, format=image_format)
                img_io.seek(0)

                document_instance.tampered_doc.save(f"tampered_{document_instance.submission_id}.{image_format.lower()}", ContentFile(img_io.read()))
                document_instance.save()

            else:
                analysis_response.data["tampered_doc_url"] = "Unsupported file format. Only PDF, PNG, JPG, and JPEG are supported."
                return Response(data={"status": True, "message": analysis_response.data}, status=200)

            # Add tampered doc URL to response
            analysis_response.data["tampered_doc_url"] = document_instance.tampered_doc.url
            return Response(data={"status": True, "message": analysis_response.data}, status=200)

        except Exception as ex:
            print(f"Error: {ex}")
            return Response(data={"status": True, "message": analysis_response.data}, status=200)


                                        